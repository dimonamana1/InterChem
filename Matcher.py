# -*- coding: utf-8 -*-
import csv
import os
import re
from difflib import SequenceMatcher
import openpyxl
import xlrd
from xlsxwriter import Workbook
from openpyxl import Workbook
import Widgets


class Matcher:
    def __init__(self, directory, reportsdirectory, rfilepath, rfilepath2):
        self.__directory = directory
        self.__reportsdirectory = reportsdirectory
        self.__rfilepath = rfilepath
        self.__rfilepath2 = rfilepath2

    __dictdb = {}  # drugs dict
    __dictallbd = {}  # garbage dict

    def __listd(self):  # files from directory list
        return os.listdir(self.__directory)

    @staticmethod
    def __csv_writer(path, data):  # write data to path
        with open(path, "a", newline='') as csv_file:
            writer = csv.writer(csv_file, delimiter=';')
            writer.writerow(data)

    @staticmethod
    def __is_digit(string):
        if string.isdigit():
            return True
        else:
            try:
                float(string)
                return True
            except ValueError:
                return False

    @staticmethod
    def __is_garbage(string):
        digit = 0
        non_digit = 0
        for i in string:
            if i.isdigit():
                digit += 1
            else:
                non_digit += 1
        if digit / (digit + non_digit) > 0.4:
            return True
        else:
            return False

    @staticmethod
    def setrfilepath(filepath, filepath2):
        Matcher.__rfilepath = filepath
        Matcher.__rfilepath2 = filepath2

    @staticmethod
    def setdirectory(directory):
        Matcher.__directory = directory

    @staticmethod
    def setreportsdirectory(directory):
        Matcher.__reportsdirectory = directory

    @staticmethod
    def setdrugsfilepath(directory):
        Matcher.__rfilepath = directory

    @staticmethod
    def setgarbagefilepath(directory):
        Matcher.__rfilepath2 = directory

    @staticmethod
    def __csv_from_excel(filename, datapath):  # read from datapath; write to filename
        try:
            if (datapath[-4:]) == "xlsx":
                wb = openpyxl.load_workbook(datapath)
                filenames = []
                sheets = wb.sheetnames
                for sheet in sheets:
                    sh = wb[str(sheet)]
                    with open(filename[:-4] + sheet + ".csv", 'tw', newline='') as f:
                        filenames.append(filename[:-4] + sheet + ".csv")
                        wr = csv.writer(f, quoting=csv.QUOTE_ALL, delimiter=";")
                        for row in sh.rows:
                            temp = []
                            for cell in row:
                                temp.append(cell.value)
                            wr.writerow(temp)
                        f.close()
                return filenames
            elif (datapath[-4:]) == ".xls":
                wb = xlrd.open_workbook(datapath, encoding_override="cp1251")
                sheetlist = wb.sheet_names()
                filenames = []
                for sheet in sheetlist:
                    sh = wb.sheet_by_name(sheet)
                    with open(filename[:-4] + sheet + ".csv", 'tw', newline='') as f:
                        filenames.append(filename[:-4] + sheet + ".csv")
                        wr = csv.writer(f, quoting=csv.QUOTE_ALL, delimiter=";")
                        for rownum in range(sh.nrows):
                            wr.writerow(sh.row_values(rownum))
                        f.close()
                return filenames
        except UnicodeDecodeError:
            print(datapath)

    def __response_handler(self, curpos, mlist):  # answers for responses
        try:
            ww = Widgets.Widgets()
            while len(mlist) < 4:
                mlist.append([0, "Error"])
            uinput1 = ww.showDialog1(curpos, Matcher.__dictdb[mlist[0][1]])
            if uinput1 == "y":
                data = [curpos, Matcher.__dictdb[mlist[0][1]]]
                Matcher.__csv_writer(self.__rfilepath, data)
                Matcher.__dictdb[curpos] = data[1]
            elif uinput1 == "garb":
                Matcher.__csv_writer(self.__rfilepath2, [curpos])
                Matcher.__dictallbd[curpos] = curpos
            else:
                dictemp = {"1": Matcher.__dictdb[mlist[1][1]],
                           "2": Matcher.__dictdb[mlist[2][1]],
                           "3": Matcher.__dictdb[mlist[3][1]]}
                uinput2 = ww.showDialog2(Matcher.__dictdb[mlist[1][1]], Matcher.__dictdb[mlist[2][1]],
                                         Matcher.__dictdb[mlist[3][1]])
                if uinput2 == "1" or uinput2 == "2" or uinput2 == "3":
                    data = [curpos, dictemp.get(uinput2)]
                    Matcher.__csv_writer(self.__rfilepath, data)
                    Matcher.__dictdb[curpos] = data[1]
                else:
                    uinput3 = ww.showDialog()
                    data = [curpos, uinput3]
                    Matcher.__csv_writer(self.__rfilepath, data)
                    Matcher.__dictdb[curpos] = data[1]
        except IndexError:
            print("list out of bonds")

    def __match(self, rpath, wpath):  # read form rpath and write to wpath with renamed drugs
        with open(rpath, "r", newline="") as file:
            reader = csv.reader(file, delimiter=';', quotechar='"')
            for row2 in reader:
                for i in range(len(row2)):
                    word = row2[i]
                    if Matcher.__is_digit(word):
                        word = word.replace('.', ',')
                        row2[i] = word
                    temp = re.sub(r"[#()./|i*®^ї'%і_\-$!~=\"]", "", word.lower())
                    curpos = temp.strip()
                    if curpos in Matcher.__dictdb:
                        row2[i] = Matcher.__dictdb[curpos]
                    elif Matcher.__is_digit(curpos) or curpos == "" or (
                            curpos in Matcher.__dictallbd) or Matcher.__is_garbage(curpos):
                        continue
                    else:
                        mlist = []
                        for k in Matcher.__dictdb:
                            a = SequenceMatcher(None, curpos, k).ratio()
                            if a > 0.6:
                                mlist.append([a, k])
                        if len(mlist) == 0:
                            Matcher.__csv_writer(self.__rfilepath2, [curpos])
                            Matcher.__dictallbd[curpos] = curpos
                        elif mlist[0][0] > 0.985:
                            row2[i] = mlist[0][1]
                            data = [curpos, mlist[0][1]]
                            Matcher.__csv_writer(self.__rfilepath, data)
                        elif len(mlist) != 0:
                            mlist.sort(reverse=True)
                            Matcher.__response_handler(self, curpos, mlist)
                            if curpos in Matcher.__dictdb:
                                row2[i] = Matcher.__dictdb[curpos]
                Matcher.__csv_writer(wpath, row2)

    def __exelwriter(self, csvfile):  # creates same named as csvfile xlsx and write data to it
        wb = Workbook(self.__reportsdirectory + csvfile[:-4] + '.xlsx')
        ws = wb.create_sheet("Sheet")
        with open(csvfile, 'rt', newline="") as f:
            reader = csv.reader(f, delimiter=';')
            for subarray in reader:
                subarray.append(csvfile[:-3])
                ws.append(subarray)
        wb.save(self.__reportsdirectory + csvfile[:-3] + '.xlsx')

    def rename_drugs(self):  # unites some methods:
        with open(self.__rfilepath, "r", newline="") as file:  # create drugs dict
            reader = csv.reader(file, delimiter=';', skipinitialspace=True)
            for row1 in reader:
                self.__dictdb[row1[0]] = row1[1]
        with open(self.__rfilepath2, "r", newline="") as file:  # create garbage dict
            reader = csv.reader(file, delimiter=';', skipinitialspace=True)
            for row2 in reader:
                self.__dictallbd[row2[0]] = row2[0]
        errors = []
        for file in self.__listd():
            try:
                filenames = Matcher.__csv_from_excel(file,
                                                     self.__directory + file)  # for files from directory creates csv
                # copy
                for filenm in filenames:  # for files from created list creates result files with renamed drugs
                    filenameres = (filenm[:-3] if filenm[:-3] == "xls" else filenm[:-4]) + ".csv"
                    with open(filenameres, 'tw', newline=''):  # and writes result to xslx
                        try:
                            self.__match(filenm, filenameres)
                            self.__exelwriter(filenameres)
                        except Exception as q:
                            print(q)
                    os.remove(filenm)
                    os.remove(filenameres)
            except KeyError:
                errors.append(file)
        for error in errors:
            Widgets.Widgets.showError(error)

    def createbd(self, rfilepath1):  # add drugs from filepath1 to drugs dict
        with open(self.__rfilepath, "r", newline="") as file:
            reader = csv.reader(file, delimiter=';', skipinitialspace=True)
            for row1 in reader:
                self.__dictdb[row1[0]] = row1[1]
        with open(rfilepath1, "r", newline="") as file:
            reader = csv.reader(file, delimiter=';', skipinitialspace=True)
            for row2 in reader:
                temp = re.sub(r"[#()./|i*®^ї'%і_\-$!~=\"]", "", row2[0].lower())
                curpos = temp.strip()
                if curpos in self.__dictdb:
                    continue
                else:
                    mlist = []
                    for i in self.__dictdb:
                        mlist.append([SequenceMatcher(None, curpos, i).ratio(), i])
                    mlist.sort(reverse=True)
                    Matcher.__response_handler(self, curpos, mlist)
